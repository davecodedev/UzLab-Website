from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.worksheet.datavalidation import DataValidation

FONT = "Arial"
HEADER_FILL = PatternFill("solid", start_color="16181D", end_color="16181D")
HEADER_FONT = Font(name=FONT, bold=True, color="FFFFFF")
EXAMPLE_FILL = PatternFill("solid", start_color="F6F7F9", end_color="F6F7F9")
EXAMPLE_FONT = Font(name=FONT, italic=True, color="6B7280")
BODY_FONT = Font(name=FONT)
TITLE_FONT = Font(name=FONT, bold=True, size=14)
NOTE_FONT = Font(name=FONT, italic=True, color="6B7280")

wb = Workbook()

# ---------------------------------------------------------------- Instructions
ws = wb.active
ws.title = "Instructions"
ws.column_dimensions["A"].width = 100
lines = [
    ("How to use this file", TITLE_FONT),
    ("", BODY_FONT),
    ("There are three tabs below: Publications, News, and Membership Types.", BODY_FONT),
    ("Each has a header row (do not change the column names) and one example", BODY_FONT),
    ("row in gray italics — delete the example row before sending this back,", BODY_FONT),
    ("or just add your real rows below it.", BODY_FONT),
    ("", BODY_FONT),
    ("Columns marked (required) must be filled in for every row.", BODY_FONT),
    ("Columns marked (optional) can be left blank.", BODY_FONT),
    ("", BODY_FONT),
    ("Publication / News \"Published date\": leave blank to save as a draft", BODY_FONT),
    ("(saved but not shown on the public site) — useful if it's not ready yet.", BODY_FONT),
    ("Use YYYY-MM-DD format, e.g. 2026-03-15.", BODY_FONT),
    ("", BODY_FONT),
    ("Publication \"Tags\": comma-separated, e.g.  accreditation, testing, labs", BODY_FONT),
    ("", BODY_FONT),
    ("Membership \"Price\": whole amount in the given currency, e.g. 500000", BODY_FONT),
    ("(not cents) for 500,000 UZS.", BODY_FONT),
    ("", BODY_FONT),
    ("File attachments (PDFs etc. for publications) aren't wired up on the site", BODY_FONT),
    ("yet — for now, put the full text in \"Body text\". We'll follow up about", BODY_FONT),
    ("attaching original documents once that's ready.", BODY_FONT),
    ("", BODY_FONT),
    ("When you're done, send this file back as-is.", BODY_FONT),
]
for i, (text, font) in enumerate(lines, start=1):
    cell = ws.cell(row=i, column=1, value=text)
    cell.font = font

# ---------------------------------------------------------------- Publications
ws = wb.create_sheet("Publications")
headers = [
    "Title (required)",
    "Category (required)",
    "Summary (required)",
    "Body text (required)",
    "Language (optional, default uz)",
    "Tags (optional, comma-separated)",
    "Author (optional)",
    "Published date (optional, YYYY-MM-DD, blank = draft)",
]
widths = [28, 20, 40, 60, 18, 28, 20, 28]
for col, (h, w) in enumerate(zip(headers, widths), start=1):
    c = ws.cell(row=1, column=col, value=h)
    c.font = HEADER_FONT
    c.fill = HEADER_FILL
    c.alignment = Alignment(wrap_text=True, vertical="center")
    ws.column_dimensions[c.column_letter].width = w
ws.row_dimensions[1].height = 30
ws.freeze_panes = "A2"

example = [
    "ISO 17025 Accreditation Cookbook",
    "COOKBOOK",
    "Step-by-step guide to preparing a lab accreditation submission.",
    "Full text of the guide goes here — as much as you like.",
    "uz",
    "accreditation, labs",
    "A. Karimov",
    "2026-03-15",
]
for col, val in enumerate(example, start=1):
    c = ws.cell(row=2, column=col, value=val)
    c.font = EXAMPLE_FONT
    c.fill = EXAMPLE_FILL

dv_category = DataValidation(
    type="list", formula1='"COOKBOOK,LEGISLATIVE,INTERNATIONAL_LITERATURE"', allow_blank=False
)
ws.add_data_validation(dv_category)
dv_category.add("B2:B500")

for row in range(2, 501):
    for col in range(1, len(headers) + 1):
        if row == 2:
            continue
        ws.cell(row=row, column=col).font = BODY_FONT

# ---------------------------------------------------------------- News
ws = wb.create_sheet("News")
headers = [
    "Title (required)",
    "Summary (required)",
    "Body text (required)",
    "Published date (optional, YYYY-MM-DD, blank = draft)",
]
widths = [28, 40, 60, 28]
for col, (h, w) in enumerate(zip(headers, widths), start=1):
    c = ws.cell(row=1, column=col, value=h)
    c.font = HEADER_FONT
    c.fill = HEADER_FILL
    c.alignment = Alignment(wrap_text=True, vertical="center")
    ws.column_dimensions[c.column_letter].width = w
ws.row_dimensions[1].height = 30
ws.freeze_panes = "A2"

example = [
    "UzLab Annual Conference 2026",
    "Save the date for our annual conference.",
    "Full announcement text goes here.",
    "2026-04-01",
]
for col, val in enumerate(example, start=1):
    c = ws.cell(row=2, column=col, value=val)
    c.font = EXAMPLE_FONT
    c.fill = EXAMPLE_FILL

for row in range(3, 501):
    for col in range(1, len(headers) + 1):
        ws.cell(row=row, column=col).font = BODY_FONT

# ---------------------------------------------------------------- Membership Types
ws = wb.create_sheet("Membership Types")
headers = [
    "Name (required)",
    "Description (required)",
    "Price (required, whole amount)",
    "Currency (optional, default UZS)",
    "Duration in days (required, e.g. 365)",
    "Active (optional, TRUE/FALSE, default TRUE)",
]
widths = [24, 45, 22, 18, 20, 18]
for col, (h, w) in enumerate(zip(headers, widths), start=1):
    c = ws.cell(row=1, column=col, value=h)
    c.font = HEADER_FONT
    c.fill = HEADER_FILL
    c.alignment = Alignment(wrap_text=True, vertical="center")
    ws.column_dimensions[c.column_letter].width = w
ws.row_dimensions[1].height = 30
ws.freeze_panes = "A2"

example = ["Individual", "Standard individual membership.", 500000, "UZS", 365, "TRUE"]
for col, val in enumerate(example, start=1):
    c = ws.cell(row=2, column=col, value=val)
    c.font = EXAMPLE_FONT
    c.fill = EXAMPLE_FILL

dv_bool = DataValidation(type="list", formula1='"TRUE,FALSE"', allow_blank=True)
ws.add_data_validation(dv_bool)
dv_bool.add("F2:F500")

for row in range(3, 501):
    for col in range(1, len(headers) + 1):
        ws.cell(row=row, column=col).font = BODY_FONT

wb.save("UzLab_Content_Template.xlsx")
print("saved")
